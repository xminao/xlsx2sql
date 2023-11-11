'''
-*- coding: utf-8 -*-
@File: xlsx2sql.py
@Author: Minhao
@Desc: 用于根据DB_Design模板文件生成对应建表语句
@Time: 2023/11/04

'''

import openpyxl
import sys
import os
import time

# 生成的脚本头内容
SCRIPT_CONTENT = ''
SCRIPT_AUTHOR = ''
SCRIPT_VERSION = 'v1.0'

# 模板文件
TEMPLATE_DB_DESIGN = '.\\template\\[TEMPLATE]_DB_Design.xlsx'
TEMPLATE_SQL = '.\\template\\template_sql.sql'
TEMPLATE_SCRIPT = '.\\template\\template_script.sql'

# 输入输出目录
INPUT_PATH = '.\\input'
OUTPUT_PATH = '.\\output'

# 输入文件后缀，用于校验
INPUT_EXTENDSION = 'DB_Design.xlsx'

# 输出日志(待完成)
LOG_PATH = '.\\logs'
LOG_FORMAT = '%Y%m%d-%H%M%S'

"""
    校验文件的格式与模板是否一致：
    1. 列名以及顺序是否一致
    file_path: 输入文件的名（包括路径）
"""
def check_file_format(file_path: str):
    # 确认输入文件存在 Table_List页 和 TableColumn页
    input_wb = openpyxl.load_workbook(file_path)
    if ('Table_List' not in input_wb.sheetnames) or ('TableColumn' not in input_wb.sheetnames):
        raise Exception(f"请确认 {file_path} 存在 Table_List 和 TableColumn Sheet页.")

    # 验证Table_List列名以及顺序
    # 1. 读取模板 DB_Design.xlsx 的 Table_List页 列名
    template_table_ws = openpyxl.load_workbook(TEMPLATE_DB_DESIGN)['Table_List']
    template_table_ws_title = []
    for row in template_table_ws.iter_rows(min_row=2, max_row=2, values_only=True):
        for col in row:
            template_table_ws_title.append(col)

    # 2. 读取输入文件的 Table_List页 列名
    input_table_ws = openpyxl.load_workbook(file_path)['Table_List']
    input_table_ws_title = []
    for row in input_table_ws.iter_rows(min_row=2, max_row=2, values_only=True):
        for col in row:
            input_table_ws_title.append(col)

    # 3. 判断输入文件列名以及顺序是否与模板一致
    if template_table_ws_title != input_table_ws_title:
        raise Exception(f"输入文件 {file_path} 的 Table_List 页格式与模板不一致.")

    # 验证TableColumn列名以及顺序
    # 1. 读取模板 DB_Design.xlsx 的 TableColumn页 列名
    template_column_ws = openpyxl.load_workbook(TEMPLATE_DB_DESIGN)['TableColumn']
    template_column_ws_title = []
    for row in template_column_ws.iter_rows(min_row=2, max_row=2, values_only=True):
        for col in row:
            template_column_ws_title.append(col)

    # 2. 读取输入文件的 TableColumn页 列名
    input_column_ws = openpyxl.load_workbook(file_path)['TableColumn']
    input_column_ws_title = []
    for row in input_column_ws.iter_rows(min_row=2, max_row=2, values_only=True):
        for col in row:
            input_column_ws_title.append(col)

    # 3. 判断输入文件列名以及顺序是否与模板一致
    if template_column_ws_title != input_column_ws_title:
        raise Exception(f"输入文件 {file_path} 的 TableColumn 页格式与模板不一致.")


"""
    检验输入文件内容：
    1. Table_List的数据有效位有没有非法的（非数值）
    2. Table_List有效的数据关键列是不是空（数据库名称，数据库，schema，表名）
    3. Table_List有效的数据表名在TableColumn表有没有有效的合法数据字段
    file_path: 输入文件的名（包括路径）
"""
def check_file_content(file_path):
    # 读取DB_Design_TableList页
    file_ws = openpyxl.load_workbook(file_path)['Table_List']

"""
    根据生成的schema字典数据集，生成SQL脚本
    schema_dict: 传入的字典数据集
"""
def generate_sql_script(schema_dict: dict, table_comment: dict, filename: str):
    # 读取SQL模板文件
    with open(TEMPLATE_SQL, 'r') as template_file:
        sql_template = template_file.read()

    # 生成SQL脚本
    sql_script = ''
    
    #{'dbo': {'table1': [['col1', ['int', 'Y', 'N', 'COMMENT']], ['col2', ['int', 'Y', 'N', 'COMMENT']]]}}
    # 遍历dbo字典 
    for sch_k, sch_v in schema_dict.items():
        # 遍历该dbo下的table字典
        for tb_k, tb_v in sch_v.items():
            script = '' # 这个表的创建脚本
            content = '' # 字段相关的内容
            pk_content = 'PRIMARY KEY (' # 主键内容
            pk_count = 0
            # 遍历该table下的所有col数组
            for index, col in enumerate(tb_v):
                nullable = '' # NOT NULL约束
                if index != len(tb_v) - 1:
                    nullable = ',' if col[1][1] == 'Y' else ' NOT NULL,'
                else:
                    nullable = '&end' if col[1][1] == 'Y' else ' NOT NULL&end'

                comment = '' # 注释
                if col[1][3] is not None:
                    comment = '' if col[1][3] == '' else ' -- ' + col[1][3]

                # 追加content
                if index != len(tb_v) - 1:
                    content += col[0] + ' ' + col[1][0] + nullable + comment + '\n\t'
                else:
                    content += col[0] + ' ' + col[1][0] + nullable + comment

                if col[1][2] == 'Y': # 是否是主键
                    pk_count += 1
                    pk_content += col[0] + ', '
            pk_content += ')'
            script = sql_template.replace('[schema_user]', '[' + sch_k + ']' or '')\
                                    .replace('[table_name]', '[' + tb_k + ']' or '')\
                                    .replace('[content]', content)\
            # 如果有主键就替换内容
            if pk_count != 0:
                script = script.replace('[primary_key]', pk_content)\
                                .replace(', )', ')')\
                                .replace('&end', ',')
            else:
                script = script.replace('[primary_key]', '')\
                                .replace('\t\n);', ');')\
                                .replace('&end', '')
            print(f'表: {tb_k}')
            if table_comment[tb_k] is not None:
                script = script.replace('[table_comment]', table_comment[tb_k])

            sql_script += script + '\n\n'

    # 输出的SQL文件名
    output_sql_file = OUTPUT_PATH + '.\\' + os.path.splitext(filename)[0] + '.sql'

    # 将创建表的SQL脚本保存到文件
    # with open(output_sql_file, 'w', encoding='utf-8') as sql_file:
    #     sql_file.write(sql_script)

    # print(f"SQL脚本已生成并保存到 {output_sql_file}")

    # 生成一个SCRIPT（带大注释头）
    # 读取SQL模板文件
    with open(TEMPLATE_SCRIPT, 'r', encoding='utf-8') as template_script:
        script_template = template_script.read()

    sql_scripts = ''
    sql_scripts = script_template.replace('[scripts]', sql_script)\
                                    .replace('[scripts_created_date]', time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))\
                                    .replace('[scripts_content]', SCRIPT_CONTENT)\
                                    .replace('[scripts_author]', SCRIPT_AUTHOR)\
                                    .replace('[scripts_version]', SCRIPT_VERSION)
                    
    with open(output_sql_file, 'w', encoding='utf-8') as sql_file:
        sql_file.write(sql_scripts)

    print(f"SQL脚本已生成并保存到 {output_sql_file}")
    return


"""
    主函数，执行整个过程
"""
def main():
    # 使用 os 模块列出目录中的文件
    for filename in os.listdir(INPUT_PATH):
        file_path = os.path.join(INPUT_PATH, filename)  # 构建完整的文件路径
        print(file_path)
        print(filename)

        # 检查路径是否是一个文件，且以指定格式结尾
        if os.path.isfile(file_path) and file_path.endswith(INPUT_EXTENDSION):
            # 校验文件格式是否与模板一致
            try:
                check_file_format(file_path)
            except Exception as ex:
                print(f"File Format Error: {ex}")
                sys.exit(1)

            # 读取输入文件数据
            input_file_ws = openpyxl.load_workbook(file_path)['TableColumn']
            data = []
            illegal_data = []
            invalid_data = []
            for row_num, row in enumerate(input_file_ws.iter_rows(min_row=3, values_only=True), start=3):
                schema_user, table_name, column_name, data_type, nullable, db_pk, column_comments, business_pk, column_description, sample_data, StatusReason, active_status = row
                if all(cell is None for cell in row):
                    pass
                if type(active_status) is int: 
                    # 判断数据有效位是否有值
                    if schema_user is not None and table_name is not None and column_name is not None and nullable is not None:
                        # 关键列不能为空，schema, table_name, column_name, data_type, nullable
                        if nullable != 'Y' and nullable != 'N':
                            # 非空列nullable值要合法：Y / N
                            illegal_data.append((row_num, 'nullalbe列值非法'))
                        elif db_pk is not None and db_pk != 'Y' and db_pk != 'N':
                            # 主键列db_pk值要合法：Y / N / None
                            illegal_data.append((row_num, 'db_pk列值非法'))
                        # 数据有效位表示为非0
                        elif active_status != 0:
                            data.append((schema_user, table_name, column_name, data_type, nullable, db_pk, column_comments, business_pk, column_description, sample_data, StatusReason, active_status))
                        else:
                            invalid_data.append(row_num)
                    else:
                        illegal_data.append((row_num, '关键列有空值'))
                else:
                    illegal_data.append((row_num, 'active_status列值非法'))
            print('有效表数据：')
            for row in data:
                print(row)
            print('非法数据行：')
            for row in illegal_data:
                print(f"- {row}")
            print('失效数据行：')
            print(invalid_data)

            # 读取表中schema, table_name, column(包含属性)
            # {'dbo': {'table1': [['col1', ['Y', 'N', 'COMMENT']], ['col2', ['Y', 'N', 'COMMENT']]]}}
            schema_dict = {}
            for row in data:
                # if schmea not in dict, init
                schema = row[0]
                table_dict = schema_dict.get(schema, {})
                table = row[1]
                col_list = table_dict.get(table, [])
                col_attr = [row[3], row[4], row[5], row[6]]
                col = [row[2], col_attr]
                # update col_list
                col_list.append(col)
                # update table_dict
                table_dict[table] = col_list
                # update schema_dict
                schema_dict[schema] = table_dict
            print("读取进数据结构的数据:")
            for sch_k, sch_v in schema_dict.items():
                print(f'schema: {sch_k}')
                for tb_k, tb_v in sch_v.items():
                    print(f'- table: {tb_k}')
                    for col in tb_v:
                        print(f'\t- col: {col}')

            # 读取Table_List，获取表名对应注释
            table_comment = {}
            table_list_ws = openpyxl.load_workbook(file_path)['Table_List']
            for row in table_list_ws.iter_rows(min_row=3, values_only=True):
                no, server_info, db_name, schema_user, classification, table_name, table_description, table_comments_detail, data_scope, num_rows, StatusReason, active_status = row
                if active_status == 1:
                    table_comment[table_name] = table_description
            print('表注释：')
            for tc_k, tc_v in table_comment.items():
                print(f"table:{tc_k}, comment:{tc_v}")

            # 生成SQL脚本
            generate_sql_script(schema_dict, table_comment, filename)

if __name__ == "__main__":
    main()